from LxmlSoup import LxmlSoup
import requests
from openpyxl import load_workbook



html = requests.get('https://sportpitanie.netlify.app/basket').text  # получаем html код сайта
soup = LxmlSoup(html)  # создаём экземпляр класса LxmlSoup

links = soup.find_all('span', class_='item-product-name')  # получаем список ссылок и наименований
for i, link in enumerate(links):
    name = link.text()  # извлекаем наименование из блока со ссылкой
    price = soup.find_all("span", class_="item-price")[i].text()  # извлекаем цену
    print(i)
    print(f"Name - {name}\n")
    print(f"Price - {price}\n")

    fn = 'parser.xlsx'
    wb = load_workbook(fn)
    ws = wb['Лист1']
    ws['A1'] = 'Название'
    ws['B1'] = 'Цена'
    for y in range(i) :
        ws[f'A{i+2}'] = name
        ws[f'B{i+2}'] = price
        wb.save(fn)
        wb.close()
